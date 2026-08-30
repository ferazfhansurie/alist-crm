import hashlib
import json
import re
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import frappe
from frappe.utils import get_datetime, now_datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


CANONICAL_SHEETS = [
	"Leads Apr 26 FB",
	"Leads Apr 26 TT",
	"Leads May 26 FB",
	"Leads May 26 TT",
	"Leads May - boss",
	"Leads Jun 26 FB",
	"Leads Jun 26 TT",
	"Talent Jun 26 FB",
	"Founder Series 96",
	"Leads June- Boss",
	"Leads July 26 FB",
	"Leads July 26 TT",
	"Founder Series July",
	"Talent July 26 FB",
	"Leads July - Boss",
	"Leads Aug 26 FB",
	"Leads Aug 26 TT",
	"Founder Series Aug 26",
	"Leads Aug - Boss",
	"Leads Past Client",
]

REPORT_SHEETS = {
	"Daily Report April 26": 4,
	"Daily Report May 26": 5,
	"Daily Report Jun 26": 6,
	"Daily Report July 26": 7,
	"Daily Report Aug 26": 8,
}

MONTHS = {
	"JANUARY": 1,
	"FEBRUARY": 2,
	"MARCH": 3,
	"APRIL": 4,
	"MAY": 5,
	"JUNE": 6,
	"JULY": 7,
	"AUGUST": 8,
	"SEPTEMBER": 9,
	"OCTOBER": 10,
	"NOVEMBER": 11,
	"DECEMBER": 12,
}

OWNERS = {"TIKA": "Tika", "IZZY": "Izzy", "AIMAN": "Aiman", "FATIN": "Fatin", "FAREEZ": "Fareez"}

ALIASES = {
	"lead_datetime": {"date time", "date & time", "date", "created time", "created_time", "start timing"},
	"external_id": {"id", "lead id", "lead_id"},
	"name": {"nama", "name", "full name", "nama penuh", "client"},
	"phone": {"no tel", "no telefon", "phone", "phone number", "hp number", "contact number", "mobile no"},
	"email": {"email", "email address"},
	"company": {"company", "company name", "nama syarikat", "business name"},
	"pic": {"pic", "owner", "to pass to", "pass to"},
	"status": {"status", "lead status"},
	"annual_sales": {"sales tahunan", "annual sales", "sales 2025", "sales setahun"},
	"monthly_sales": {"sales bulanan", "monthly sales", "sales sebulan"},
	"business_type": {"business type", "jenis bisnes", "type of business"},
	"service_required": {"service required", "service needed", "servis diperlukan"},
	"pax": {"pax kehadiran", "pax", "attendance"},
	"ad_id": {"ad id", "ad_id"},
	"ad_name": {"ad name", "ad_name"},
	"adset_id": {"adset id", "adset_id"},
	"adset_name": {"adset name", "adset_name"},
	"campaign_id": {"campaign id", "campaign_id"},
	"campaign_name": {"campaign name", "campaign_name"},
	"form_id": {"form id", "form_id"},
	"form_name": {"form name", "form_name"},
}

ACTIVITY_PATTERNS = [
	("DAH CONTACT", "Contact", "Contacted"),
	("CALL, NO PICKUP", "Call", "No Pickup"),
	("CALL NO PICKUP", "Call", "No Pickup"),
	("CALL, PICKUP", "Call", "Pickup"),
	("CALL PICKUP", "Call", "Pickup"),
	("CALL TAK ANGKAT", "Call", "No Pickup"),
	("CALL ANGKAT", "Call", "Pickup"),
	("NO WHATSAPP", "WhatsApp", "No WhatsApp"),
	("WHATSAPP REPLY", "WhatsApp", "Replied"),
	("MEETING SET", "Meeting", "Set"),
	("DAH SET MEETING", "Meeting", "Set"),
	("DAH MEETING", "Meeting", "Completed"),
	("PROPOSAL REQUESTED", "Proposal", "Proposal Requested"),
	("PROPOSAL SENT", "Proposal", "Proposal Requested"),
	("SIGNED CLIENT", "Conversion", "Signed Client"),
	("SIGNED CONTRACT", "Conversion", "Signed Client"),
	("NO RESPONSE", "Contact", "No Response"),
	("NON-QUALITY", "Qualification", "Non-Quality"),
	("NON QUALITY", "Qualification", "Non-Quality"),
	("BAD LEAD", "Qualification", "Bad Lead"),
	("REDUNDANT", "Qualification", "Redundant"),
	("CONFIRMED", "Event", "Confirmed"),
	("DECLINE", "Event", "Declined"),
]


def _header(value) -> str:
	text = str(value or "").strip().lower().replace("_", " ")
	return re.sub(r"\s+", " ", text)


def _text(value) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _phone(value) -> str:
	text = _text(value)
	if not text:
		return ""
	digits = re.sub(r"\D", "", text)
	if len(digits) < 7 or len(digits) > 15:
		return ""
	if digits.startswith("60"):
		return f"+{digits}"
	if digits.startswith("0"):
		return f"+60{digits[1:]}"
	return f"+60{digits}" if 8 <= len(digits) <= 11 else f"+{digits}"


def _email(value) -> str:
	return _text(value).lower()


def _date(value):
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, datetime.min.time())
	text = _text(value)
	if not text:
		return None
	for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S%z", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%m/%d/%Y %H:%M"):
		try:
			parsed = datetime.strptime(text, fmt)
			return parsed.replace(tzinfo=None)
		except ValueError:
			continue
	try:
		return get_datetime(text)
	except Exception:
		return None


def _channel(sheet: str) -> tuple[str, str]:
	name = sheet.lower()
	if "past client" in name:
		return "Past Client", "Past Client"
	if "founder" in name:
		return "Founder Series", "Founder Series"
	if "talent" in name:
		return "Talent", "Talent"
	if "boss" in name:
		return "Boss / Manual", "Normal Lead"
	if "tt" in name:
		return "TikTok", "Normal Lead"
	return "Meta", "Normal Lead"


def _status_column(formula_sheet, header_row: int, headers: dict[int, str]) -> int | None:
	for col, header in headers.items():
		if header in ALIASES["status"]:
			return col
	for validation in formula_sheet.data_validations.dataValidation:
		formula = _text(validation.formula1).upper()
		if "DAH CONTACT" not in formula and "CONFIRMED" not in formula:
			continue
		for cell_range in validation.sqref.ranges:
			min_col, min_row, _, _ = range_boundaries(str(cell_range))
			if min_row <= header_row + 2:
				return min_col
	return None


def _find_header_row(sheet) -> int:
	best_row, best_score = 1, -1
	known = set().union(*ALIASES.values())
	for row in range(1, min(sheet.max_row, 6) + 1):
		score = sum(1 for cell in sheet[row] if _header(cell.value) in known)
		if score > best_score:
			best_row, best_score = row, score
	return best_row


def _column_map(value_sheet, formula_sheet, header_row: int) -> dict[str, int | None]:
	headers = {cell.column: _header(cell.value) for cell in value_sheet[header_row] if _header(cell.value)}
	mapping = {}
	for field, aliases in ALIASES.items():
		mapping[field] = next((col for col, header in headers.items() if header in aliases), None)
	mapping["status"] = mapping["status"] or _status_column(formula_sheet, header_row, headers)

	if not mapping["pic"]:
		owner_hits = Counter()
		for col in range(1, value_sheet.max_column + 1):
			for row in range(header_row + 1, min(value_sheet.max_row, header_row + 150) + 1):
				if _text(value_sheet.cell(row, col).value).upper() in OWNERS:
					owner_hits[col] += 1
		mapping["pic"] = owner_hits.most_common(1)[0][0] if owner_hits else None

	status_col = mapping["status"]
	mapping["remark"] = status_col + 1 if status_col and status_col < value_sheet.max_column else None
	return mapping


def _value(sheet, row: int, col: int | None):
	return sheet.cell(row, col).value if col else None


def _number(value) -> float:
	try:
		return float(value or 0)
	except (TypeError, ValueError):
		return 0


def _reporting_data(path: str) -> dict:
	values = load_workbook(Path(path), data_only=True, read_only=False)
	year = 2026
	monthly = []
	closings = []
	daily_rows = []

	summary = values["Summary Leads"]
	for row in range(2, min(summary.max_row, 14) + 1):
		month_number = MONTHS.get(_text(summary.cell(row, 1).value).upper())
		if month_number not in REPORT_SHEETS.values():
			continue
		monthly.append(
			{
				"month": f"{year}-{month_number:02d}",
				"leads": int(_number(summary.cell(row, 2).value)),
				"meetings": int(_number(summary.cell(row, 3).value)),
				"closed": int(_number(summary.cell(row, 4).value)),
				"closed_amount": _number(summary.cell(row, 5).value),
				"spend": _number(summary.cell(row, 7).value),
			}
		)

	closing_month = None
	for row in range(19, min(summary.max_row, 100) + 1):
		month_number = MONTHS.get(_text(summary.cell(row, 1).value).upper())
		if month_number:
			closing_month = month_number
		client = _text(summary.cell(row, 2).value)
		amount_value = summary.cell(row, 3).value
		if not closing_month or not client or not isinstance(amount_value, (int, float)):
			continue
		source = _text(summary.cell(row, 4).value)
		source_month = next(
			(f"{year}-{number:02d}" for label, number in MONTHS.items() if label in source.upper()),
			None,
		)
		channel = "Website" if "WEBSITE" in source.upper() else "Founder Series" if "OPEN DAY" in source.upper() else "Meta"
		closings.append(
			{
				"deal": f"xlsx-closing-{row}",
				"client": client,
				"amount": _number(amount_value),
				"closed_on": date(year, closing_month, monthrange(year, closing_month)[1]).isoformat(),
				"source_month": source_month,
				"channel": channel,
			}
		)

	for sheet_name, month_number in REPORT_SHEETS.items():
		sheet = values[sheet_name]
		header_row = next(
			(row for row in range(1, min(sheet.max_row, 25) + 1) if _header(sheet.cell(row, 1).value) == "date"),
			None,
		)
		if not header_row:
			continue
		headers = {cell.column: _header(cell.value) for cell in sheet[header_row] if _header(cell.value)}
		lead_col = next((col for col, label in headers.items() if label == "total meta leads"), None)
		meeting_col = next((col for col, label in headers.items() if label == "total meetings"), None)
		spend_col = next((col for col, label in headers.items() if label == "meta ad spend"), None)
		remark_col = next((col for col, label in headers.items() if label == "remark"), None)
		awareness_col = next((col for col, label in headers.items() if label == "awareness"), None)
		meta_leads = meta_meetings = 0
		meta_spend = 0.0
		first_data_row = header_row + 2
		last_data_row = first_data_row + monthrange(year, month_number)[1]
		for row in range(first_data_row, last_data_row):
			report_date = _date(sheet.cell(row, 1).value)
			if not report_date or report_date.year != year or report_date.month != month_number:
				continue
			reported_leads = int(_number(_value(sheet, row, lead_col)))
			reported_meetings = int(_number(_value(sheet, row, meeting_col)))
			lead_spend = _number(_value(sheet, row, spend_col))
			meta_leads += reported_leads
			meta_meetings += reported_meetings
			meta_spend += lead_spend
			daily_rows.append(
				{
					"date": report_date.date().isoformat(),
					"channel": "Meta",
					"reported_leads": reported_leads,
					"reported_meetings": reported_meetings,
					"monthly_adjustment": 0,
					"lead_spend": lead_spend,
					"awareness_spend": _number(_value(sheet, row, awareness_col)),
					"remark": _text(_value(sheet, row, remark_col)) or None,
				}
			)

		target = next((item for item in monthly if item["month"] == f"{year}-{month_number:02d}"), None)
		if not target:
			continue
		remaining_leads = target["leads"] - meta_leads
		remaining_meetings = target["meetings"] - meta_meetings
		remaining_spend = target["spend"] - meta_spend
		for row in range(6, min(header_row, 12)):
			label = _text(sheet.cell(row, 1).value)
			channel = {"TIKTOK": "TikTok", "GOOGLE": "Google", "FOUNDER SERIES": "Founder Series"}.get(label.upper())
			if not channel:
				continue
			reported_leads = max(min(int(_number(sheet.cell(row, 2).value)), remaining_leads), 0)
			reported_meetings = max(min(int(_number(sheet.cell(row, 9).value)), remaining_meetings), 0)
			lead_spend = max(min(_number(sheet.cell(row, 12).value), remaining_spend), 0)
			remaining_leads -= reported_leads
			remaining_meetings -= reported_meetings
			remaining_spend -= lead_spend
			if not any((reported_leads, reported_meetings, lead_spend)):
				continue
			daily_rows.append(
				{
					"date": date(year, month_number, 1).isoformat(),
					"channel": channel,
					"reported_leads": reported_leads,
					"reported_meetings": reported_meetings,
					"monthly_adjustment": 1,
					"lead_spend": lead_spend,
					"awareness_spend": 0,
					"remark": "Workbook monthly channel total",
				}
			)

	return {"monthly": monthly, "closings": closings, "daily": daily_rows}


def _import_reporting(path: str) -> dict:
	reporting = _reporting_data(path)
	settings = frappe.get_single("A-List Settings")
	settings.historical_summary_json = json.dumps(reporting["monthly"], default=str)
	settings.historical_closings_json = json.dumps(reporting["closings"], default=str)
	settings.save(ignore_permissions=True)
	created = updated = 0
	for row in reporting["daily"]:
		name = frappe.db.exists("A-List Daily Marketing", {"date": row["date"], "channel": row["channel"]})
		if name:
			doc = frappe.get_doc("A-List Daily Marketing", name)
			updated += 1
		else:
			doc = frappe.new_doc("A-List Daily Marketing")
			created += 1
		for field, value in row.items():
			doc.set(field, value)
		doc.save(ignore_permissions=True)
	return {
		"monthly_snapshots": len(reporting["monthly"]),
		"historical_closings": len(reporting["closings"]),
		"daily_marketing_created": created,
		"daily_marketing_updated": updated,
	}


def _lifecycle(raw_status: str) -> str:
	status = raw_status.upper()
	if "REDUNDANT" in status:
		return "Duplicate"
	if any(word in status for word in ("BAD LEAD", "NON-QUALITY", "NON QUALITY", "DECLINE")):
		return "Disqualified"
	if any(word in status for word in ("SIGNED CLIENT", "SIGNED CONTRACT", "PROPOSAL REQUESTED", "PROPOSAL SENT")):
		return "Converted"
	if "DAH MEETING" in status:
		return "Meeting Done"
	if "MEETING SET" in status or "DAH SET MEETING" in status:
		return "Meeting Set"
	if status:
		return "Contacted"
	return "New"


def _last_outcome(raw_status: str) -> str | None:
	status = raw_status.upper()
	for phrase, _, outcome in reversed(ACTIVITY_PATTERNS):
		if phrase in status and outcome in {
			"Pickup", "No Pickup", "Replied", "No WhatsApp", "Set", "Completed",
			"Proposal Requested", "No Response", "Non-Quality", "Bad Lead", "Redundant",
		}:
			return {"Set": "Meeting Set", "Completed": "Meeting Done"}.get(outcome, outcome)
	return None


def _activities(raw_status: str, occurred_at, owner: str | None) -> list[dict]:
	status = raw_status.upper()
	rows = []
	for phrase, activity_type, outcome in ACTIVITY_PATTERNS:
		if phrase in status and not any(row["outcome"] == outcome and row["activity_type"] == activity_type for row in rows):
			rows.append({"activity_type": activity_type, "outcome": outcome, "occurred_at": occurred_at})
	for raw_owner, label in OWNERS.items():
		if f"PASS TO {raw_owner}" in status or f"{raw_owner} AMBIL" in status or f"{raw_owner} AMBIK" in status:
			rows.append({"activity_type": "Reassignment", "outcome": "Reassigned", "occurred_at": occurred_at, "note": f"Passed to {label}"})
	return rows


def _import_key(channel: str, external_id: str, email: str, phone: str, lead_date, sheet: str, row: int) -> str:
	# Excel treats every source row as a lead submission, including repeat
	# contacts and repeat platform IDs. Keep that row identity in Frappe too;
	# contact-level deduplication would silently erase valid submissions.
	basis = f"{sheet}|row|{row}"
	return "xlsx:" + hashlib.sha256(basis.encode()).hexdigest()[:32]


def read_workbook(path: str) -> tuple[list[dict], dict]:
	path_obj = Path(path)
	if not path_obj.exists():
		frappe.throw(f"Workbook not found: {path}")
	values = load_workbook(path_obj, data_only=True, read_only=False)
	formulas = load_workbook(path_obj, data_only=False, read_only=False)
	records = []
	report = {"workbook": path_obj.name, "sheets": {}, "warnings": []}

	for sheet_name in CANONICAL_SHEETS:
		if sheet_name not in values.sheetnames:
			report["warnings"].append(f"Missing expected sheet: {sheet_name}")
			continue
		value_sheet = values[sheet_name]
		formula_sheet = formulas[sheet_name]
		header_row = _find_header_row(value_sheet)
		columns = _column_map(value_sheet, formula_sheet, header_row)
		channel, stream = _channel(sheet_name)
		sheet_count = 0

		for row in range(header_row + 1, value_sheet.max_row + 1):
			name = _text(_value(value_sheet, row, columns["name"]))
			raw_phone = _text(_value(value_sheet, row, columns["phone"]))
			phone = _phone(raw_phone)
			email = _email(_value(value_sheet, row, columns["email"]))
			company = _text(_value(value_sheet, row, columns["company"]))
			if not any((name, raw_phone, email, company)):
				continue
			lead_date = _date(_value(value_sheet, row, columns["lead_datetime"]))
			external_id = _text(_value(value_sheet, row, columns["external_id"]))
			owner_raw = _text(_value(value_sheet, row, columns["pic"])).upper()
			owner = OWNERS.get(owner_raw, owner_raw.title() if owner_raw else None)
			raw_status = _text(_value(value_sheet, row, columns["status"]))
			key = _import_key(channel, external_id, email, phone, lead_date, sheet_name, row)
			remark = _text(_value(value_sheet, row, columns["remark"]))
			if raw_phone and not phone:
				remark = " · ".join(filter(None, [remark, f"Original contact: {raw_phone}"]))
			records.append(
				{
					"lead": {
						"first_name": name or company or phone or email or raw_phone or f"Lead {row}",
						"email": email or None,
						"mobile_no": phone or None,
						"organization": company or None,
						"status": _lifecycle(raw_status),
						"source": channel,
						"alist_lead_datetime": lead_date,
						"alist_channel": channel,
						"alist_stream": stream,
						"alist_annual_sales_band": _text(_value(value_sheet, row, columns["annual_sales"])) or None,
						"alist_monthly_sales_text": _text(_value(value_sheet, row, columns["monthly_sales"])) or None,
						"alist_business_type": _text(_value(value_sheet, row, columns["business_type"])) or None,
						"alist_service_required": _text(_value(value_sheet, row, columns["service_required"])) or None,
						"alist_pax": _value(value_sheet, row, columns["pax"]),
						"alist_pic_name": owner,
						"alist_last_outcome": _last_outcome(raw_status),
						"alist_event_outcome": "Confirmed" if "CONFIRMED" in raw_status.upper() else "Declined" if "DECLINE" in raw_status.upper() else None,
						"alist_remark": remark or None,
						"alist_original_status": raw_status or None,
						"alist_external_lead_id": external_id or None,
						"alist_ad_id": _text(_value(value_sheet, row, columns["ad_id"])) or None,
						"alist_ad_name": _text(_value(value_sheet, row, columns["ad_name"])) or None,
						"alist_adset_id": _text(_value(value_sheet, row, columns["adset_id"])) or None,
						"alist_adset_name": _text(_value(value_sheet, row, columns["adset_name"])) or None,
						"alist_campaign_id": _text(_value(value_sheet, row, columns["campaign_id"])) or None,
						"alist_campaign_name": _text(_value(value_sheet, row, columns["campaign_name"])) or None,
						"alist_form_id": _text(_value(value_sheet, row, columns["form_id"])) or None,
						"alist_form_name": _text(_value(value_sheet, row, columns["form_name"])) or None,
						"alist_source_tab": sheet_name,
						"alist_source_row": row,
						"alist_import_key": key,
					},
					"activities": _activities(raw_status, lead_date or now_datetime(), owner),
				}
			)
			sheet_count += 1

		report["sheets"][sheet_name] = {
			"rows": sheet_count,
			"header_row": header_row,
			"mapped_columns": {field: col for field, col in columns.items() if col},
		}
	return records, report


def dry_run(path: str) -> dict:
	records, report = read_workbook(path)
	keys = [record["lead"]["alist_import_key"] for record in records]
	duplicate_keys = {key for key, count in Counter(keys).items() if count > 1}
	existing = set()
	for offset in range(0, len(keys), 500):
		existing.update(
			row.alist_import_key
			for row in frappe.get_all(
				"CRM Lead",
				filters={"alist_import_key": ["in", keys[offset : offset + 500]]},
				fields=["alist_import_key"],
				limit_page_length=0,
			)
		)
	unique_new = {key for key in keys if key not in existing}
	seen_new = set()
	new_activity_count = 0
	for record in reversed(records):
		key = record["lead"]["alist_import_key"]
		if key not in unique_new or key in seen_new:
			continue
		seen_new.add(key)
		new_activity_count += len(record["activities"])
	report.update(
		{
			"rows_read": len(records),
			"unique_import_keys": len(set(keys)),
			"duplicate_rows_in_workbook": len(records) - len(set(keys)),
			"duplicate_keys": len(duplicate_keys),
			"already_imported": len(existing),
			"new_leads": len(unique_new),
			"new_activities": new_activity_count,
			"by_channel": dict(Counter(record["lead"]["alist_channel"] for record in records)),
			"by_status": dict(Counter(record["lead"]["status"] for record in records)),
		}
	)
	return report


def import_workbook(path: str) -> dict:
	records, _ = read_workbook(path)
	created = 0
	activities_created = 0
	skipped = 0
	seen = set()
	# Duplicate contacts can reappear in a later monthly tab. Import the latest
	# occurrence so the CRM starts with the workbook's current owner and status.
	for record in reversed(records):
		lead_data = record["lead"]
		key = lead_data["alist_import_key"]
		if key in seen or frappe.db.exists("CRM Lead", {"alist_import_key": key}):
			skipped += 1
			continue
		seen.add(key)
		lead = frappe.get_doc({"doctype": "CRM Lead", **lead_data}).insert(ignore_permissions=True)
		created += 1
		for activity_data in record["activities"]:
			frappe.get_doc(
				{
					"doctype": "A-List Lead Activity",
					"lead": lead.name,
					"actor": frappe.session.user,
					"source_label": "Workbook import",
					**activity_data,
				}
			).insert(ignore_permissions=True)
			activities_created += 1
	reporting = _import_reporting(path)
	frappe.db.commit()
	return {
		"created_leads": created,
		"created_activities": activities_created,
		"skipped": skipped,
		**reporting,
	}


def dry_run_json(path: str) -> str:
	return json.dumps(dry_run(path), default=str, indent=2)
